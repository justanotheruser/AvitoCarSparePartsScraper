import json
import logging
import os
import sys

import dacite
import openpyxl
import pika

sys.path.append(os.path.dirname(__file__) + "/..")
from common.logger import setup_logger
from common.config import read_config
from common.queues import FAILED_QUERIES_QUEUE
from common.types import SearchQuery


def main():
    setup_logger('failed_queries_handler.log')
    cfg = read_config()
    if not cfg:
        return

    failed_queries_file = cfg.output.failed_queries_file
    logging.info(f'Failed queries file: {failed_queries_file}')

    wb = openpyxl.Workbook()
    sheet = wb.active
    for i, column_header in enumerate(SearchQuery.csv_header()):
        sheet.cell(row=1, column=i+1).value = column_header
    wb.save(failed_queries_file)

    connection = pika.BlockingConnection(pika.ConnectionParameters(host='localhost'))
    channel = connection.channel()
    channel.queue_declare(queue=FAILED_QUERIES_QUEUE)

    def read_query(ch, method, properties, body):
        json_query = json.loads(body)
        logging.info(f'Received query {json_query}')
        query: SearchQuery = dacite.from_dict(data_class=SearchQuery, data=json_query)
        wb = openpyxl.load_workbook(failed_queries_file)
        sheet = wb.active
        row_number = sheet.max_row + 1
        sheet.cell(row=row_number, column=1).value = query.car_brand
        sheet.cell(row=row_number, column=2).value = query.car_model
        sheet.cell(row=row_number, column=3).value = query.spare_part
        wb.save(failed_queries_file)
        ch.basic_ack(delivery_tag=method.delivery_tag)

    channel.basic_consume(queue=FAILED_QUERIES_QUEUE, on_message_callback=read_query)
    logging.info('Waiting for messages')
    channel.start_consuming()
    connection.close()


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        logging.error(e)

