import json
import logging
import os
import sys

import dacite
import pika

sys.path.append(os.path.dirname(__file__) + "/..")
from common.logger import setup_logger
from common.config import read_config
from common.queues import SCRAPED_ITEMS_QUEUE
from common.types import ScrapedItem
import openpyxl


def main():
    setup_logger('result_writer.log')
    cfg = read_config()
    if not cfg:
        return

    result_file = cfg.output.result_file
    logging.info(f'Result file: {result_file}')
    if not os.path.exists(result_file):
        wb = openpyxl.Workbook()
        sheet = wb.active
        for i, column_header in enumerate(ScrapedItem.csv_header()):
            sheet.cell(row=1, column=i+1).value = column_header
        wb.save(result_file)

    connection = pika.BlockingConnection(pika.ConnectionParameters(host='localhost'))
    channel = connection.channel()
    channel.queue_declare(queue=SCRAPED_ITEMS_QUEUE)

    def read_item(ch, method, properties, body):
        item = json.loads(body)
        logging.info(f'Received item {item}')
        item = dacite.from_dict(data_class=ScrapedItem, data=item)
        wb = openpyxl.load_workbook(result_file)
        sheet = wb.active
        row_number = sheet.max_row + 1
        for i, value in enumerate(item.as_csv_columns()):
            if type(value) not in (int, float, str):
                value = str(value)
            sheet.cell(row=row_number, column=i+1).value = value
        wb.save(result_file)
        ch.basic_ack(delivery_tag=method.delivery_tag)

    channel.basic_consume(queue=SCRAPED_ITEMS_QUEUE, on_message_callback=read_item)
    logging.info('Waiting for messages')
    channel.start_consuming()
    connection.close()


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print('Interrupted')
        try:
            sys.exit(0)
        except SystemExit:
            os._exit(0)
